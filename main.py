#!/usr/bin/python
# -*- coding: GBK -*-
import copy
import threading
import time
import tkinter
from tkinter import Button, Label, Entry
from collections import defaultdict
from tkinter.filedialog import askdirectory

import openpyxl
import os

_SUFFIX_LIT = ['.xls', '.xlsx']
_OUTPUT_NAME = 'result'
_OUTPUT_SUFFIX = '.xlsx'
_OUTPUT_FOLDER = 'output'

# "C:\\Users\\Logon\\Desktop\\"


def get_merged_cells_value(merged, sheet, row_index, col_index):
    for (_, scol), (_, srow), (_, ecol), (_, erow) in merged:
        if srow <= row_index <= erow:
            if scol <= col_index <= ecol:
                cell_value = sheet.cell(srow, scol).value
                # print("get", row_index, col_index, srow, scol, cell_value)
                return cell_value
                break
    return None


def obtain_excel_files(folder_path):
    """ get all excel files from in the specified folder with the suffix xls or xlsx
    :param folder_path: Absolute path of folder containing excel file
    :return: excel file list
    """
    excel_list = []
    walk = os.walk(folder_path)
    for cur_path, dir_list, file_list in walk:
        print(dir_list)
        for file_name in file_list:
            if os.path.splitext(file_name)[-1] in _SUFFIX_LIT:
                if _OUTPUT_FOLDER not in cur_path:
                    print(os.path.dirname(os.path.join(cur_path, file_name)))
                    excel_list.append(os.path.join(cur_path, file_name))
    return excel_list


def transfer_rows(pyxl_row):
    row_list = []
    for cell in pyxl_row:
        row_list.append(cell.value)
    print(row_list)
    return row_list


class ExcelFilter:
    def __init__(self, thread, output_path):
        self.__thread = thread
        self.output_path = output_path
        file_name = _OUTPUT_NAME + "_" + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + _OUTPUT_SUFFIX
        self.output_file = os.path.join(self.output_path, file_name)
        print(self.output_file)

    def write_result(self, key_row):
        if not os.path.exists(self.output_path):
            os.mkdir(self.output_path)
        # f = open(self.output_file, 'w')
        # f.close()
        if not os.path.exists(self.output_file):
            out_wb = openpyxl.Workbook()
        else:
            out_wb = openpyxl.load_workbook(self.output_file)
        for key, value_rows in key_row.items():
            if not(self.__thread.get_thread_status()):
                break
            print(key)
            try:
                out_sheet = out_wb[key]
            except KeyError:
                out_sheet = out_wb.create_sheet(key)
            for value_row in value_rows:
                out_sheet.append(value_row)
        out_wb.save(self.output_file)

    def obtain_row_contain_key(self, sheet, keys):
        """ obtain row that contain key
        :param sheet: sheet
        :param keys: the list of key
        :return: key_row
        """
        key_row = defaultdict(list)
        merged = sheet.merged_cells
        row_num = sheet.max_row
        col_num = sheet.max_column
        for row_idx in range(1, row_num):
            if not(self.__thread.get_thread_status()):
                break
            for col_idx in range(1, col_num):
                cell_value = sheet.cell(row_idx, col_idx).value
                if not(cell_value is None or cell_value == '') and cell_value in keys:
                    row_list = []
                    for cur_col in range(1, col_num):
                        cur_value = sheet.cell(row_idx, cur_col).value
                        if cur_value is None or cur_value == '':
                            # print('merged', row, cur_col)
                            cur_value = get_merged_cells_value(merged, sheet, row_idx, cur_col)
                        row_list.append(cur_value)
                    print(row_list)
                    key_row[cell_value].append(tuple(row_list))
        return key_row

    def filter_main(self, in_path, keys):
        filter_path = in_path
        filter_keys = keys
        excel_files = []
        if os.path.isdir(filter_path):
            excel_files = obtain_excel_files(filter_path)
        elif os.path.isfile(filter_path):
            excel_name = os.path.basename(filter_path)
            if os.path.splitext(excel_name)[-1] in _SUFFIX_LIT:
                excel_files.append(filter_path)
        else:
            print("it's not a folder or file.")
        for excel_file in excel_files:
            if not(self.__thread.get_thread_status()):
                break
            print("process file:" + excel_file)
            self.process_file(excel_file, filter_keys)

    def process_file(self, file, keys):
        """
        :param file:
        :param keys:
        :return:
        """
        file_wb = openpyxl.load_workbook(file)
        print("load file done")
        sheets = file_wb.sheetnames
        for per_sheet in sheets:
            file_sheet = file_wb[per_sheet]
            if not(self.__thread.get_thread_status()):
                break
            print("process sheet:"+file_sheet.title)
            self.write_result(self.obtain_row_contain_key(file_sheet, keys))


class FilterThread(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.path = ""
        self.keys = ""
        self.__is_running = True

    def set_filter_param(self, path, keys):
        self.path = path
        self.keys = keys

    def thread_terminate(self):
        self.__is_running = False

    def get_thread_status(self):
        return self.__is_running

    def run(self):
        excel_filter = ExcelFilter(self, os.path.join(self.path, _OUTPUT_FOLDER))
        excel_filter.filter_main(self.path, self.keys)
        print("Done!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print("完成了..........................................")

    # Need to get rid of the repetition


class FileWindow:
    def __init__(self, width, height):
        self.window = tkinter.Tk()
        self.width = width
        self.height = height
        self.folder_path = tkinter.StringVar()
        self.keys_list = tkinter.StringVar()
        self.__filter_thread = FilterThread()
        self.__thread_status = 0

    def select_folder(self):
        self.folder_path.set(askdirectory())

    def close_window(self):
        if self.__thread_status == 1:
            self.__filter_thread.thread_terminate()
        self.window.destroy()

    def config_window(self):
        screenwidth = self.window.winfo_screenwidth()
        screenheight = self.window.winfo_screenheight()
        size = '%dx%d+%d+%d' % (self.width, self.height, (screenwidth - self.width) / 2,
                                                                (screenheight - self.height) / 2)

        self.window.geometry(size)
        self.window.protocol('WM_DELETE_WINDOW', self.close_window)

    def set_view(self):
        Label(self.window, text='文件夹路径:').grid(row=0, column=0)
        Entry(self.window, textvariable=self.folder_path).grid(row=0, column=1)
        Button(self.window, text='选择文件夹', command=self.select_folder).grid(row=0, column=2)
        Label(self.window, text='关键字:').grid(row=1, column=0)
        Entry(self.window, textvariable=self.keys_list).grid(row=1, column=1)
        Button(self.window, text='开始过滤', command=self.start_filter).grid(row=2, column=2)
        self.window.mainloop()

    def start_filter(self):
        get_path = self.folder_path.get()
        get_keys = self.keys_list.get()
        if not(get_path == "" or get_keys == ""):
            self.__filter_thread.set_filter_param(get_path, get_keys.split())
            self.__filter_thread.start()
            self.__thread_status = 1


if __name__ == "__main__":
    if 1:
        # target = "C:\\Users\\Logon\\Desktop\\中文路径\\"
        FileWindow = FileWindow(400, 500)
        FileWindow.config_window()
        FileWindow.set_view()
    else:
        path = "E:/result_2020-02-23-22:58-49.txt"
        file = open(path, "wb+")
        file.write("123")
        file.close()

